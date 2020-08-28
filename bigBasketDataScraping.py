from selenium import webdriver
import time
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


br=webdriver.Firefox()
url="https://www.bigbasket.com/cl/foodgrains-oil-masala/?nc=nb#!page=1"
br.get(url)


br.find_element_by_xpath("/html/body/div[1]/div[1]/header/div/div/div/div/ul/li[2]/div/a").click()
time.sleep(1)
br.find_element_by_xpath("/html/body/div[1]/div[1]/header/div/div/div/div/ul/li[2]/div/div/div[2]/form/div[1]/div/div/span").click()
time.sleep(1)
br.find_element_by_xpath("/html/body/div[1]/div[1]/header/div/div/div/div/ul/li[2]/div/div/div[2]/form/div[1]/div/input[1]").send_keys("Hyderabad")
time.sleep(1)
br.find_element_by_xpath("/html/body/div[1]/div[1]/header/div/div/div/div/ul/li[2]/div/div/div[2]/form/div[1]/div/ul[1]/li/div[3]/a").click()
time.sleep(1)
br.find_element_by_xpath("/html/body/div[1]/div[1]/header/div/div/div/div/ul/li[2]/div/div/div[2]/form/div[3]/button").click()
time.sleep(1)

names=[]
mp=[]
sp=[]
qty=[]

moreDataAvailable=True
page=1
count=0
br.get(url)
while moreDataAvailable==True:
    br.execute_script("window.scrollTo(0, document.body.scrollHeight);var lenOfPage=document.body.scrollHeight;return lenOfPage;")
    try:
        moreButton=br.find_element_by_xpath("/html/body/div[1]/div[3]/product-deck/section/div[2]/div[4]/div[3]/button")
        moreButton.click()
    except:
       print("Button not found") 
    time.sleep(2)
    page+=1
    if moreDataAvailable==False:
        print("No more data available")
        break
    for i in range(count+1,count+25):
        count+=1
        if i%5!=0:
            nameXpath="/html/body/div[1]/div[3]/product-deck/section/div[2]/div[4]/div[1]/div/div/div[2]/div/div[%s]/product-template/div/div[4]/div[1]/a"%(i)
            mpXpath="/html/body/div[1]/div[3]/product-deck/section/div[2]/div[4]/div[1]/div/div/div[2]/div/div[%s]/product-template/div/div[4]/div[3]/div/div[1]/h4/span[1]/span"%(i)
            spXpath="/html/body/div[1]/div[3]/product-deck/section/div[2]/div[4]/div[1]/div/div/div[2]/div/div[%s]/product-template/div/div[4]/div[3]/div/div[1]/h4/span[2]/span"%(i)
            pXpath="/html/body/div[1]/div[3]/product-deck/section/div[2]/div[4]/div[1]/div/div/div[2]/div/div[%s]/product-template/div/div[4]/div[3]/div/div[1]/h4/span/span[2]"%(i)
            singleQtyXpath="/html/body/div[1]/div[3]/product-deck/section/div[2]/div[4]/div[1]/div/div/div[2]/div/div[%s]/product-template/div/div[4]/div[2]/div[1]/span/span[1]"%(i)
            multipleQtyXpath="/html/body/div[1]/div[3]/product-deck/section/div[2]/div[4]/div[1]/div/div/div[2]/div/div[%s]/product-template/div/div[4]/div[2]/div/span/button/span/span[1]"%(i)

            #getting name of the product, which if not available meaning end of the list
            try:
                name=br.find_element_by_xpath(nameXpath)
                names.append(name.text)
            except:
                print("element does not exist no more data available")
                moreDataAvailable=False
                break

            #getting quantity of the product if available as a single quantity
            try:
                quantity=br.find_element_by_xpath(singleQtyXpath)
                qty.append(quantity.text)
            except:
                print("single quantity does not exist")
                #try getting if quantity is available as a dropdown
                try:
                    quantity=br.find_element_by_xpath(multipleQtyXpath)
                    qty.append(quantity.text)
                except:
                    print("Quantity not found")
                    qty.append("not found")

            #getting the actual price of the product if available
            try:
                mprice=br.find_element_by_xpath(mpXpath)
                mp.append(mprice.text)
            except:
                print("marked price does not exist")
                mp.append(0)

            #getting the market or discounted price of the product if available
            try:
                sprice=br.find_element_by_xpath(spXpath)
                sp.append(sprice.text)
            except:
                print("No discounted price available")
                #getting the actual price of the product if available
                try:
                    sprice=br.find_element_by_xpath(pXpath)
                    sp.append(sprice.text)
                except:
                    print("No price value available")
                    sp.append(0)
                    continue
print(names)
print(len(names))
print(mp)
print(len(mp))
print(sp)
print(len(sp))
print(qty)
print(len(qty))

#the excel sheet does not exist and must be created
print("creating Workbook")
workbook=openpyxl.Workbook()
currentSheet=workbook.active
#Complete everything

currentSheet.cell(1,1).value="Product Name"
currentSheet.cell(1,2).value="Quantity"
currentSheet.cell(1,3).value="Marked price"
currentSheet.cell(1,4).value="Selling price"

rowCount=2
#add names, qty, mp and sp to the sheet
print("adding products to excel sheet")
for i in range(len(names)):
    currentSheet.cell(rowCount,1).value=names[i]
    currentSheet.cell(rowCount,2).value=qty[i]
    currentSheet.cell(rowCount,3).value=mp[i]
    currentSheet.cell(rowCount,4).value=sp[i]
    rowCount+=1

#make sure to chenge this to your desktop directory
workbook.save('C:\\Users\\91970\\Desktop\\bb_foodgrains_oil_masala.xlsx')
#lastly don't forget to save



